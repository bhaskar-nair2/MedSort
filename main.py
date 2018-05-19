from openpyxl import Workbook, load_workbook
import itertools
import search

wbList = load_workbook('Demand.xlsx') # Demand Sheet
wbSearch = load_workbook('SearchFile.xlsx') # Search Sheet

wbList._active_sheet_index = 0
ws = wbList.active
names = ws['C'][2:]
values = ws['R'][2:]

for a,b in itertools.zip_longest(names,values):
    if a.value is None:
        break
    for n in range(0, len(wbSearch.sheetnames)):
        wbSearch._active_sheet_index = n
        wsS = wbSearch.active
        nmc=wsS['D'][2:]
        valx=wsS['C'][2:]

        for x,y in itertools.zip_longest(nmc,valx):
            if x.value is None:
                break
            elif search.isSimilar(x.value,a.value):
            #elif SequenceMatcher(None, x.value, a.value).ratio()>6:
                #a.value=x.value
                b.value=y.value
                print("Found "+a.value+" at "+ y.value+"\n\n")

wbList.save('Demand.xlsx')

