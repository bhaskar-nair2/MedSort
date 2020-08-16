from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
import sqlite3 as sql
from re import sub
from datetime import date
import os
import sys


class ReDataMaker:
    def __init__(self, gui):
        data = gui.get()
        file = data[0]
        self.showError = gui.showError
        self.text = gui.log
        self.pgbar = gui.pgbar

        try:
            self.file = load_workbook(data[0], data_only=True)
        except (InvalidFileException, FileNotFoundError) as e:
            self.showError(
                'File Error', "No file or File Format not supported (Only supports .xlsx files)")

        self.pa_count = data[3]

    def refresh(self):
        self.pgbar.step(1)
        print('Start')
        try:
            con = sql.connect(db, isolation_level=None)
            self.cur = con.cursor()
        except sql.OperationalError as e:
            print(f'Error in init: {e}')
        self.text.insert("end", "Refresh Starts!\n")
        tbl_set = 0
        self.clear_db()
        rcTab = False
        for sheet_index in range(0, len(self.file.sheetnames)):
            self.file._active_sheet_index = sheet_index
            self.text.insert("end",
                             f"Now Inserting {self.file.sheetnames[self.file._active_sheet_index]} into Database \n")
            sheet = self.file.active

            if not sheet_index < self.pa_count[tbl_set]:
                tbl_set += 1
                self.pgbar.step(33)
                print(tables, tbl_set)
                if tbl_set == len(tables):
                    break
            if(tables[tbl_set] == 'rcSearchList'):
                rcTab = True
            contract_lst = sheet['C'][1:]
            name_lst = sheet['D'][1:]
            unit_lst = sheet['E'][1:]
            coy_lst = sheet['F'][1:]
            rate_lst = sheet['G'][1:]
            gst_lst = sheet['J'][1:]
            supplier_lst = sheet['L'][1:]
            if(rcTab):
                to_lst = sheet['M'][1:]
                fr_lst = sheet['N'][1:]
            for _ in range(len(contract_lst)):
                if contract_lst[_].value is None:
                    break
                value = [contract_lst[_].value,
                         str(name_lst[_].value).lower().strip(),
                         makeAlias(str(name_lst[_].value)),
                         unit_lst[_].value,
                         coy_lst[_].value,
                         rate_lst[_].value,
                         gst_lst[_].value,
                         str(supplier_lst[_].value).replace('\n', '')]
                if(rcTab):
                    value.append(to_lst[_].value)
                    value.append(fr_lst[_].value)
                self.insert(tables[tbl_set], value, rcTab)
        self.text.insert.put('Refresh Done!!')

    def clear_db(self):
        try:
            for _ in tables:
                que = f"Delete from {_}"
                self.cur.execute(que)
        except sql.OperationalError:
            print('Error IN deletion')

    def insert(self, tbl, values, rcTab):
        if(rcTab):
            que = f"insert into {tbl} (contract, name, alias, unit, coy, rate, gst, supplier, to_date, from_date) values (?,?,?,?,?,?,?,?,?,?)"
        else:
            que = f"insert into {tbl} (contract, name, alias, unit, coy, rate, gst, supplier) values (?,?,?,?,?,?,?,?)"
        try:
            self.cur.execute(que, values)
        except sql.OperationalError as e:
            print(e)
            if(rcTab):
                re = f"""create table {tbl} (
                            contract varchar(20),
                            name     varchar(200) not null,
                            alias    varchar(200) not null,
                            unit     varchar(20),
                            coy      varchar(30),
                            rate     int default 0,
                            gst      int default 12,
                            supplier varchar(50),
                            to_date varchar(50),
                            from_date varchar(50),
                            primary key (contract,supplier)
                            )"""
            else:
                re = f"""create table {tbl} (
                            contract varchar(20),
                            name     varchar(200) not null,
                            alias    varchar(200) not null,
                            unit     varchar(20),
                            coy      varchar(30),
                            rate     int default 0,
                            gst      int default 12,
                            supplier varchar(50),
                            primary key (contract,supplier)
                            )"""
            self.cur.execute(re)
            self.cur.execute(que, values)
        except Exception as e:
            print(self.__class__, f"Error{e}")
            pass


# Helper Functions
def clean(val, type):
    val = val.lower().strip()
    for _ in alias_ignore:
        val = val.replace(_, '').strip()
    if type == 'guess':
        for _ in guess_ignore:
            val = val.replace(_, '').strip()
    val = sub(r'[(\[].+?[)\]]', ' ', val)
    val = sub(r'[!@#$%^&*(),.?`\'"/:{}|<>+=-]', ' ', val)
    return val


def makeAlias(val):
    val = clean(val, 'alias')
    val = val.split(' ')
    val.sort()
    val = ''.join(val)
    val = ''.join(e for e in val if e.isalnum())
    return val.replace(' ', '')


def makePrimary(val):
    val = clean(val, 'guess')
    arr = val.split(' ')
    f_arr = list(filter(lambda e: len(e) > 6, arr))
    if len(f_arr) == 0:
        f_arr = f_arr = list(filter(lambda e: len(e) >= 4, arr))
    f_arr.sort(key=len, reverse=True)
    return f_arr[0:5]


def createUniqueTupleList(list):
    uniq = []
    for x in list:
        if x[0] not in (y[0] for y in uniq):
            uniq.append(x)
    return uniq


# Static Data
db = 'DB/SearchDB'

views = [{'ref': 'mrc_view', 'table': 'rc'},
         {'ref': 'gpa_view', 'table': 'gpa'},
         {'ref': 'spa_view', 'table': 'spa'}]

tables = ['gpaSearchList', 'spaSearchList', 'rcSearchList']


alias_ignore = ['tab', 'mg', 'oint', 'ml', 'ointment', 'cap', 'per',
                'mg', 'gm', 'amp', 'ampoule', 'cream', 'bott', 'bottle']

guess_ignore = ['insulin', 'collection', 'purified', 'equivalent',
                'containing', 'prefilled', 'syringe', 'closing',
                'polythene', 'envelope', 'facility', 'disposable',
                'plastic', 'sterile', 'suspension', 'inhaler', 'needles',
                'injection', 'culture', 'solution', 'combination',
                'sulphate', 'acetate', 'chloride',
                'test', 'kit', 'water', 'vial'
                ]
