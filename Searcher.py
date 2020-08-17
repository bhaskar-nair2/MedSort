from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
import sqlite3 as sql
from re import sub
from datetime import date
import os
import sys
import subprocess


class Searcher:
    def __init__(self, gui):
        self.frame = gui.childFrame.interior
        self.data = gui.get()
        self.childFr = {}
        self.item = 0
        self.butt = gui.btn
        self.progg = gui.pgbar
        self.text = gui.text
        self.maketablet = gui.maketablet
        self.remWid = gui.remWidget
        self.db = db
        self.showError = gui.showError

        self.tbl_name = 'indentList'
        self.columns = ['Indent No', 'Contract No', "Nomenclature",
                        "Unit", "Company", "Rate", "Quantity", "Amount", "GST", "Total Amount", "Supplier", 'from_date', "to_date"]

        try:
            self.file = load_workbook(self.data[0], data_only=True)
        except (InvalidFileException, FileNotFoundError) as e:
            self.showError(
                'File Error', "No file or File Format not supported (Only supports .xlsx files)")

        self.file_name = self.data[0].split('/')[-1].split('.')[0]
        self.NCol = self.data[1]
        self.VCol = self.data[2]
        self.progg["maximum"] = 100

    def orcestrator(self):
        print('Start')
        try:
            con = sql.connect(db, isolation_level=None)
            self.cur = con.cursor()
        except sql.OperationalError as e:
            self.showError('SQL ERROR', e)
            print(self.__class__, f"Error {e}")
        # DATA GEN
        self.progg.step(1)
        self.text.insert('end', "Started")
        self.data_gen()
        self.text.insert('end', "\n" + "Data Generated")
        # ADD FOUND
        self.progg.step(10)
        self.add_found_results()
        self.text.insert('end', "\n" + "Searching Done")
        # SEARCH BY PRIMARY
        self.progg.step(20)
        self.search_by_primary()
        self.text.insert('end', "\n" + "Guessing ¯\_(ツ)_/¯")
        # ADD NOT FOUND
        self.progg.step(50)
        self.add_not_found()
        self.text.insert('end', "\n" + "Adding Not Found")
        # RETURN FILE
        self.progg.step(19)
        self.ret_file()
        self.text.insert(
            'end', "\n\n" + f"Results File Created @{self.resultFilePath}")
        self.open_file()

    def data_gen(self):
        flag = False
        self.text.insert('end', f"\nSorting Starts at {date.today()} \n")
        self.clear_db()
        self.create_table()
        self.file._active_sheet_index = 0
        sheet = self.file.active
        indref_lst = sheet[self.data[1]][1:]
        name_lst = sheet[self.data[2]][1:]
        qty_lst = sheet[self.data[3]][1:]
        for _ in range(len(name_lst)):
            if _ % 50 == 0 or _ == len(name_lst):
                flag = False
            if name_lst[_].value is None:
                flag = True
                continue
            if flag == True:
                break
            value = [indref_lst[_].value, str(name_lst[_].value).lower().strip(),
                     makeAlias(str(name_lst[_].value)),
                     qty_lst[_].value,
                     ]
            self.insert(value)
        self.create_views()

    def add_found_results(self):
        today = date.today()
        self.WB = Workbook()
        try:
            for view in views:
                ws = self.WB.create_sheet(f"{view['table']}-{today}", 0)
                rs_found = self.cur.execute(
                    f"select * from {view['ref']} order by indref")
                ws.append(self.columns)
                res_list = rs_found.fetchall()
                for _ in res_list:
                    print(_)
                    ws.append(_)
        except sql.OperationalError as e:
            self.showError(
                'Write Error', "Error Writing to file. Run as administrator?")
            print(self.__class__, f"Error {e}")

    def search_by_primary(self):
        ws = self.WB.create_sheet("Guesses", 0)
        rs_not_found = self.cur.execute('select * from not_found;')
        rs_nf_list = rs_not_found.fetchall()
        for _ in rs_nf_list:
            ws.append(_)
            gs_list = []
            primArr = makePrimary(_[1])
            for prim in primArr:
                for tab in tables:
                    que = f"""select contract, name, coy, rate, gst, supplier from {tab} where name like '%{prim}%'"""
                    rs_guess = self.cur.execute(que)
                    gs_list.extend(rs_guess.fetchall())
            gs_list = list(set(gs_list))
            gs_list = createUniqueTupleList(gs_list)
            for j in gs_list:
                ws.append(j)
            ws.append(["----------------------",
                       "¯\_(ツ)_/¯", "----------------------"])

    def add_not_found(self):
        ws = self.WB.create_sheet("Not Found", 0)
        ws.append(["Indent Ref", "Name", "Quantity"])
        rs_not_found = self.cur.execute('select * from not_found;')
        for _ in rs_not_found:
            ws.append(_)

    def ret_file(self):
        # Create File
        filename = f"created{self.file_name}.xlsx"
        filepath = './outputs/' + filename
        self.resultFilePath = filepath
        self.WB.save(filepath)
        print('Done')
        return filepath

    def open_file(self):
        try:
            subprocess.run(self.resultFilePath, check=True)
        except PermissionError as e:
            print(f"Cant Open File {e}")

    def insert(self, values):
        que = f"insert into {self.tbl_name} (indref, name, alias, qty) values(?,?,?,?)"
        try:
            self.cur.execute(que, values)
        except sql.OperationalError as e:
            print(self.__class__, f"Error {e}")

    def clear_db(self):
        que = f"delete from {self.tbl_name}"
        try:
            self.cur.execute(que)
        except sql.OperationalError as e:
            print(f'Error ${e}')
            pass

    def create_table(self):
        que = f"""create table {self.tbl_name} (indref varchar(200),name varchar(200), alias varchar(200),qty int)"""
        try:
            self.cur.execute(que)
        except sql.OperationalError:
            pass

    def create_views(self):
        views = [
            {'ref': 'mrc_view', 'table': 'rc', 'extra': True},
            {'ref': 'gpa_view', 'table': 'gpa', 'extra': False},
            {'ref': 'spa_view', 'table': 'spa', 'extra': False}
        ]
        for view in views:
            cols = "i.indref, g.contract, i.name, g.unit, g.coy, g.rate, i.qty, g.rate*i.qty as amount, \
			        g.gst, (g.rate*i.qty*gst)+(g.rate*i.qty) as totalAmount,  g.supplier"
            if(view['extra'] == True):
                cols += ", g.to_date, g.from_date "
            que = f"""CREATE view {view['ref']} as
            select {cols}
            from {view['table']}SearchList g, indentList i
            WHERE g.name like "%"||i.name||"%" 
            or g.alias like "%"||i.alias||"%" 
            or i.name like "%"||g.name||"%" 
            or i.alias like "%"||g.alias||"%";"""
            try:
                self.cur.execute(que)
            except sql.OperationalError as e:
                self.drop(view['ref'], 'view')
                self.cur.execute(que)
                pass
            except e:
                print(self.__class__, f"Error: {e}")
            # Not Found
        self.not_found_view()

    def not_found_view(self):
        try:
            que = f"""CREATE view not_found as 
                select i.indref, i.name, i.qty 
                from indentList i 
                WHERE indref not in (select indref from mrc_view) 
                and indref not in (select indref from spa_view) 
                and indref not in (select indref from gpa_view);"""
            self.cur.execute(que)
        except sql.OperationalError as e:
            self.cur.execute(f"drop view not_found")
            self.cur.execute(que)

    def drop(self, name, type):
        try:
            self.cur.execute(f"drop {type} {name}")
        except sql.OperationalError as e:
            print(f"Delete Error: {e}")
            pass


# Helper Functions
def clean(val, type):
    val = val.lower().strip()
    for _ in alias_ignore:
        val = val.replace(_, '').strip()
    if type == 'guess':
        for _ in guess_ignore:
            val = val.replace(_, '').strip()
    val = sub(r'[(\[].+?[)\]]', ' ', val)
    val = sub(r'[!@#$%^&*(),.?`\'"/:{}|<>+=-]', ' ', val)
    return val


def makeAlias(val):
    val = clean(val, 'alias')
    val = val.split(' ')
    val.sort()
    val = ''.join(val)
    val = ''.join(e for e in val if e.isalnum())
    return val.replace(' ', '')


def makePrimary(val):
    val = clean(val, 'guess')
    arr = val.split(' ')
    f_arr = list(filter(lambda e: len(e) > 6, arr))
    if len(f_arr) == 0:
        f_arr = f_arr = list(filter(lambda e: len(e) >= 4, arr))
    f_arr.sort(key=len, reverse=True)
    return f_arr[0:5]


def createUniqueTupleList(list):
    uniq = []
    for x in list:
        if x[0] not in (y[0] for y in uniq):
            uniq.append(x)
    return uniq


# Static Data
db = 'DB/SearchDB'

views = [{'ref': 'mrc_view', 'table': 'rc'},
         {'ref': 'gpa_view', 'table': 'gpa'},
         {'ref': 'spa_view', 'table': 'spa'}]

tables = ['gpaSearchList', 'spaSearchList', 'rcSearchList']


alias_ignore = ['tab', 'mg', 'oint', 'ml', 'ointment', 'cap', 'per',
                'mg', 'gm', 'amp', 'ampoule', 'cream', 'bott', 'bottle']

guess_ignore = ['insulin', 'collection', 'purified', 'equivalent',
                'containing', 'prefilled', 'syringe', 'closing',
                'polythene', 'envelope', 'facility', 'disposable',
                'plastic', 'sterile', 'suspension', 'inhaler', 'needles',
                'injection', 'culture', 'solution', 'combination',
                'sulphate', 'acetate', 'chloride',
                'test', 'kit', 'water', 'vial'
                ]
