# Imports
from tkinter import *
from tkinter import ttk, Text, messagebox
from openpyxl import Workbook, load_workbook
import itertools
import re
import threading

# GUI Initiators
root = Tk()
root.title('Medical Sorter')
mainframe = ttk.Frame(root, padding="10 10 12 12")
mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
mainframe.columnconfigure(0, weight=1)
mainframe.rowconfigure(0, weight=1)

# Global VAriables
Indfile = StringVar(root, value='Demand.xlsx')
SrchFile = StringVar(root, value='SearchFile.xlsx')
IndenNameCol = StringVar(root, value='C')
SrchNameCol = StringVar(root, value='D')
NewValCol = StringVar(root, value='R')
ValCol = StringVar(root, value='C')
TotalEn = StringVar(root,value=271)
brute = IntVar()
fnd=0

trash = open('trash.txt', 'r').read().split("\n")
contraVals = [['tab', 'tabs', 'inj', 'bottle', 'syp', 'bot', 'bott', 'cap', 'drops', 'needles', 'ointment'],
              ['sodium', 'chloride', 'fluoride', 'phosphate']]


# Status=(root,value='Demand.xlsx')

# Classes
class ValFound(Exception):
    pass


# Functions

def contradict(a, b):
    for _ in a:
        for i in contraVals:
            if _ in i:
                for h in b:
                    if h in i and h != _:
                        return True  # Yes it does Contradict
                    if h == _:
                        return False  # No Contradiction
                    return True  # Contradicts
    for _ in b:
        for i in contraVals:
            if _ in i:
                for h in a:
                    if h in i and h != _:
                        return True
                    if h == _:
                        return False
                    return True
    return False


def remove(val):
    return list(set(val) - set(trash))


def isSimilar(v1, v2, prog, brute):
    if v1.lower() == v2.lower():
        return True
    else:
        a = re.findall(r"[\w]+", v1.lower())
        b = re.findall(r"[\w]+", v2.lower())
        if contradict(a, b):
            return False  # Stop cause it contradicts
        a = remove(a)
        b = remove(b)
        if a == b:
            return True
        else:
            if brute == 0:
                for _ in a[1:len(a) - 2]:
                    if _.isalpha():
                        for p in b:
                            if _ == p and len(_) > 5:
                                d = messagebox.askquestion('Please Help(' + str(prog) + "/"+str(TotalEn.get())+")","Is\n " + v1.upper() + "\nsimilar to\n" + v2.upper() + "\n Similar: " + _ + " : " + p + "\n [y/n]: ")
                                if d == 'yes':
                                    return True

    return False


def search():
    thread = threading.Thread(pgbar.start(),args=(None,))
    thread.start()
    # text.insert('end', Indfile.get()+"\n")
    wbList = load_workbook(Indfile.get())  # Demand Sheet
    wbSearch = load_workbook(SrchFile.get())  # Search Sheet
    prog = 0
    wbList._active_sheet_index = 0
    ws = wbList.active
    names = ws[IndenNameCol.get()][2:]
    values = ws[NewValCol.get()][2:]

    for a, b in itertools.zip_longest(names, values):
        if a.value is None:
            break
        for n in range(0, len(wbSearch.sheetnames)):
            wbSearch._active_sheet_index = n
            wsS = wbSearch.active
            nmc = wsS[SrchNameCol.get()][2:]
            valx = wsS[str(ValCol.get())][2:]
            for x, y in itertools.zip_longest(nmc, valx):
                if x.value is None:
                    break
                elif isSimilar(x.value, a.value, prog, brute.get()):
                    b.value = y.value
                    text.insert('end', "Found " + a.value + " at " + y.value + "\n\n")
                    fnd=+1
                else:
                    continue
                break
        prog += 1
    # pgbar.setV
    # pgbar.update()
    text.insert('end', "Process Done!! Found: "+str(fnd)+" out of "+str(TotalEn.get())+" \n Please Wait, Saving..")
    wbList.template=False
    wbList.save('Demand.xlsx')
    messagebox.showinfo("Done", "Operation Done")



ttk.Entry(mainframe, width=70, textvariable=Indfile).grid(column=2, row=1, sticky=(W, E), columnspan=3, pady="10")
ttk.Label(mainframe, text="Indent File ").grid(column=1, row=1, sticky=W)

ttk.Entry(mainframe, width=70, textvariable=SrchFile).grid(column=2, row=2, sticky=(W, E), columnspan=3, pady="10")
ttk.Label(mainframe, text="Search File ").grid(column=1, row=2, sticky=W)

# ttk.Separator(mainframe, orient=HORIZONTAL).grid(column=1,row=8,columnspan=4)

ttk.Entry(mainframe, width=7, textvariable=IndenNameCol).grid(column=2, row=3, sticky=(W, E), columnspan=1, pady="10")
ttk.Label(mainframe, text="Indent: Name Column ").grid(column=1, row=3, sticky=W)

ttk.Entry(mainframe, width=7, textvariable=SrchNameCol).grid(column=4, row=3, sticky=(W, E), columnspan=1, pady="10")
ttk.Label(mainframe, text="Search: Name Column ").grid(column=3, row=3, sticky=E)

ttk.Entry(mainframe, width=7, textvariable=NewValCol).grid(column=2, row=4, sticky=(W, E), pady="10")
ttk.Label(mainframe, text="Indent: Edit Column ").grid(column=1, row=4, sticky=W)

ttk.Entry(mainframe, width=7, textvariable=TotalEn).grid(column=2, row=5, sticky=(W, E), pady="10")
ttk.Label(mainframe, text="Total Values").grid(column=1, row=5, sticky=W)

C2 = ttk.Checkbutton(mainframe, text="Human Check?", variable=brute, onvalue=0, offvalue=1)
C2.grid(column=4, row=6, sticky=E)

ttk.Entry(mainframe, width=7, textvariable=ValCol).grid(column=4, row=4, sticky=(W, E), pady="10")
ttk.Label(mainframe, text="Indent: Value Column ").grid(column=3, row=4, sticky=E)

# ttk.Label(mainframe, textvariable=meters).grid(column=2, row=2, sticky=(W, E))
text = Text(mainframe, width=40, height=10)
text.grid(column=1, row=7, sticky=(W, E), columnspan=4)

pgbar = ttk.Progressbar(mainframe, orient="horizontal", mode="indeterminate")
pgbar.grid(column=1, row=8, sticky=(W, E), columnspan=4)

ttk.Button(mainframe, text="Start", command=search).grid(column=2, row=6, sticky=N, columnspan=2)



root.mainloop()
